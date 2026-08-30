#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Turn a raw IBKR trade dump into the weekly trade workbook.

Reads the JSON that get_account_trades returns, folds the fills into exit
events, derives the fields IBKR does not send (premium, entry price, return on
premium), and writes an .xlsx, a .csv and a summary JSON.

An option that expires worthless comes back as a sell at price 0 with
realized_pnl 0 — the premium is gone and the broker books nothing. Those rows
are charged the full cost of the lots that died, so the week's result is the
week's result. Sells at price 0 against nothing open are a different animal, a
platform-restart artefact, and are dropped.

Everything that can be checked is checked: the filter chain is reconciled
against the kept rows, and the cash-flow gap that IBKR's realized_pnl hides is
reported so the caller can quote it rather than discover it later.

Usage:
    python build_week.py TRADES_JSON OUTDIR [options]

Options:
    --week-start YYYY-MM-DD   First trading day to keep (default: the Monday of
                              the last full week present in the data).
    --tickers A,B,C           Keep only these underlyings.
    --drop-carry SYM[,SYM]    Drop exits of these symbols that closed a position
                              held overnight into the following day.
    --carry-day YYYY-MM-DD    The day the carried positions were closed on;
                              required with --drop-carry.
    --keep-stock              Keep trades in the shares themselves.
    --keep-buys               Keep BUY legs (one row per fill, not per exit).
    --tz-offset -4            Hours from UTC to the exchange session (default -4,
                              US Eastern in daylight time; use -5 in winter).
"""
import argparse
import collections
import csv
import datetime
import json
import os
import sys

HE_DAY = {0: "שני", 1: "שלישי", 2: "רביעי", 3: "חמישי", 4: "שישי", 5: "שבת", 6: "ראשון"}
MONEY = '$#,##0;($#,##0);-'
MONEY2 = '$#,##0.00;($#,##0.00);-'
PCT = '0.0%;(0.0%);-'


# ---------------------------------------------------------------- loading ---

def load_trades(path):
    """Accept either {"trades": [...]} or a bare list."""
    with open(path, encoding="utf-8") as fh:
        blob = json.load(fh)
    trades = blob.get("trades") if isinstance(blob, dict) else blob
    if not isinstance(trades, list):
        sys.exit("could not find a trade list in {}".format(path))
    for t in trades:
        for field in ("symbol", "sec_type", "side", "size", "price", "trade_time"):
            if field not in t:
                sys.exit("trade missing '{}': {}".format(field, t))
    return trades


def session_time(iso, tz_offset):
    utc = datetime.datetime.strptime(iso, "%Y-%m-%dT%H:%M:%SZ")
    return utc.replace(tzinfo=datetime.timezone.utc).astimezone(
        datetime.timezone(datetime.timedelta(hours=tz_offset)))


def session_day(dt, expiry=False):
    """Futures trade in an evening session that belongs to the next day.

    An expiry is the exception. IBKR stamps the write-off hours after the close —
    22:25 on the Wednesday, 22:16 on the Friday — and rolling those forward would
    date Wednesday's expiry to Thursday and Friday's to the following Monday, out of
    the week entirely. An option that expired on Friday expired on Friday.
    """
    day = dt.date()
    if expiry:
        return day
    if dt.hour >= 18:
        day += datetime.timedelta(days=1)
    while day.weekday() >= 5:
        day += datetime.timedelta(days=1)
    return day


# --------------------------------------------------------------- filtering ---

def build(trades, holds, args):
    """Apply the filter chain, keeping a running audit of what each step cost."""
    tz = args.tz_offset
    audit = [("סה\"כ עסקאות שנמשכו מ-IBKR", len(trades),
              sum(t.get("realized_pnl", 0) for t in trades))]

    def drop(rows, keep_fn, label):
        kept, gone = [], []
        for t in rows:
            (kept if keep_fn(t) else gone).append(t)
        if gone:
            audit.append((label, -len(gone), -sum(t.get("realized_pnl", 0) for t in gone)))
        return kept

    rows = trades
    if args.week_start:
        start = datetime.date.fromisoformat(args.week_start)
        end = start + datetime.timedelta(days=4)
        rows = drop(rows, lambda t: start <= session_day(session_time(t["trade_time"], tz),
                                                         t["price"] == 0) <= end,
                    "(-) מחוץ לשבוע {} עד {}".format(start, end))
    if not args.keep_stock:
        rows = drop(rows, lambda t: t["sec_type"] != "STK", "(-) עסקאות במניות עצמן")
    # Expiry rows are kept. IBKR closes a worthless option with a synthetic SELL at
    # price 0 and stamps realized_pnl 0 on it, but the whole premium is gone — the
    # feed simply never books the loss. to_exits() charges it back from the cost of
    # the lots that died, so an expiry reads as the -100% it actually was.
    def closed_something(t):
        dt = session_time(t["trade_time"], tz)
        info = holds.get((t["symbol"], dt.strftime("%Y-%m-%d %H:%M"), t["price"])) or {}
        return info.get("qty", 0) > 0
    rows = drop(rows, lambda t: t["price"] != 0 or closed_something(t),
                "(-) מכירות ב-0 ללא פוזיציה פתוחה — שאריות ריסטרט")
    if any(t["price"] == 0 for t in rows):
        audit.append(("(=) שורות פקיעה — נזקפות כהפסד מלא", 0, 0.0))
    if not args.keep_buys:
        rows = drop(rows, lambda t: t["side"] == "SELL", "(-) רגלי BUY")
    if args.tickers:
        keep = set(args.tickers)
        rows = drop(rows, lambda t: t["symbol"] in keep, "(-) טיקרים שאינם ברשימה שנמסרה")
    if args.drop_carry:
        carry, day = set(args.drop_carry), datetime.date.fromisoformat(args.carry_day)
        rows = drop(rows,
                    lambda t: not (t["symbol"] in carry
                                   and session_time(t["trade_time"], tz).date() == day),
                    "(-) פוזיציות שהועברו ללילה ({})".format(", ".join(sorted(carry))))
    return rows, audit


def holding_times(all_trades, tz):
    """How long each closing fill had been held, by first-in-first-out lot matching.

    Runs over the *unfiltered* trade list on purpose: the lots being closed were
    opened by BUY fills that the filters drop, so matching against the filtered set
    would leave most exits with no entry to measure from.

    Longs and shorts are kept in separate books, and which one a fill touches is not
    inferred from a running position — the feed says so directly. IBKR stamps
    realized_pnl on whichever leg closed something, so a BUY carrying a result is a
    buy-to-cover and a BUY carrying none opens a long, no matter what the position
    looks like from inside a seven-day window. That distinction is the whole game
    here. MU began the week already long, so its opening sells look like shorts and
    every later buy looks like a cover; read that way its Monday exits inherit the
    wrong cost basis. The restart artefacts are the mirror image — a sell against
    nothing, covered the next morning — and reading *those* as fresh longs jams
    phantom lots at the head of the queue, which is what priced the MSTR lot that
    expired on Friday at $1,440 against the $2,160 actually paid for it.

    A sell with no long to match is a position opened before the feed begins. There
    is no cost basis for it anywhere in the data, so it yields no exit here; its
    result is real and lands in the P&L, but its premium is unknowable.

    Returns {(symbol, exit_iso_minute, price): {"hold", "premium", "qty", "mult"}}
    where qty is how much of the fill closed a lot the feed can price. On a price-0
    row that is the whole test: qty 0 means nothing was open and the row is a restart
    artefact, not an expiry. The multiplier comes along because an expiry has no
    price to derive one from.
    """
    longs = collections.defaultdict(collections.deque)   # [qty, opened, cost, mult]
    shorts = collections.defaultdict(int)
    acc = collections.defaultdict(lambda: [0.0, 0, 0.0, 0.0])  # qty*min, qty, premium, qty*mult
    for t in sorted(all_trades, key=lambda x: x["trade_time"]):
        sym, qty, px = t["symbol"], t["size"], t["price"]
        dt = session_time(t["trade_time"], tz)
        mult = (t.get("net_amount", 0) / (qty * px)) if (qty and px) else 100
        if t["side"] == "BUY":
            left = qty
            if t.get("realized_pnl"):          # a result on a buy means it covered
                take = min(left, shorts[sym])
                shorts[sym] -= take
                left -= take
            if left:
                longs[sym].append([left, dt, px * mult, mult])
            continue
        key, left, book = (sym, dt.strftime("%Y-%m-%d %H:%M"), px), qty, longs[sym]
        while left > 0 and book:
            lot = book[0]
            take = min(left, lot[0])
            acc[key][0] += take * (dt - lot[1]).total_seconds() / 60.0
            acc[key][1] += take
            acc[key][2] += take * lot[2]
            acc[key][3] += take * lot[3]
            lot[0] -= take
            left -= take
            if lot[0] <= 0:
                book.popleft()
        shorts[sym] += left                    # opened before the feed, or an artefact
    return {k: {"hold": v[0] / v[1], "premium": v[2], "qty": v[1],
                "mult": round(v[3] / v[1])}
            for k, v in acc.items() if v[1]}


def to_exits(rows, tz, holds=None):
    """Fills at the same symbol, minute and price are one exit."""
    holds = holds or {}
    merged = collections.defaultdict(
        lambda: {"sz": 0, "amt": 0.0, "pnl": 0.0, "com": 0.0, "st": None})
    for t in rows:
        dt = session_time(t["trade_time"], tz)
        m = merged[(t["symbol"], dt.date(), dt.strftime("%H:%M"), t["price"])]
        m["sz"] += t["size"]
        m["amt"] += t.get("net_amount", 0.0)
        m["pnl"] += t.get("realized_pnl", 0.0)
        m["com"] += t.get("commission", 0.0)
        m["st"] = t["sec_type"]

    exits = []
    for (sym, day, hhmm, px), m in sorted(merged.items(), key=lambda kv: (kv[0][1], kv[0][2], kv[0][0])):
        # IBKR sends proceeds and net P&L; premium and entry follow from them.
        info = holds.get((sym, "{} {}".format(day, hhmm), px)) or {}
        hold = info.get("hold")
        if px == 0:
            # Nothing came back: the premium of the lots that died is the loss. Size
            # the row by what the book actually closed, not by the fill quantity —
            # the restart oversold some names, and those extra contracts were never
            # open. NU sold 390 at zero against 340 held.
            if not info.get("qty"):
                continue                      # restart artefact, dropped in build()
            m["sz"] = info["qty"]
            premium = info["premium"]
            m["pnl"] = -premium
            mult = info.get("mult") or 0    # no price here to derive it from
        else:
            mult = round(m["amt"] / (m["sz"] * px)) if m["sz"] and px else 0
            premium = m["amt"] - m["pnl"]
        exits.append(dict(
            date=day, day=HE_DAY[day.weekday()], time=hhmm, sym=sym, st=m["st"],
            qty=m["sz"], px=px, mult=mult, amt=round(m["amt"], 2),
            pnl=round(m["pnl"], 2), premium=round(premium, 2),
            entry=round(premium / (m["sz"] * mult), 4) if mult else "",
            expired=(px == 0),
            ret=round(m["pnl"] / premium, 6) if premium else "",
            hold=round(hold, 1) if hold is not None else "",
            com=round(m["com"], 2)))
    return exits


HOLD_BUCKETS = [(0, 5, "עד 5 דק׳"), (5, 30, "5–30 דק׳"), (30, 120, "30–120 דק׳"),
                (120, 480, "2–8 שעות"), (480, float("inf"), "מעל לילה")]


def hold_analysis(exits):
    """Group results by how long the position was held.

    Worth reading with the concentration check beside it: one outsized winner in a
    bucket can invert its sign, so the summary also reports each bucket without its
    largest contributor.
    """
    out = []
    for lo, hi, label in HOLD_BUCKETS:
        g = [e for e in exits
             if isinstance(e["hold"], (int, float))
             and (e["hold"] <= hi if lo == 0 else lo < e["hold"] <= hi)]
        if not g:
            continue
        # Outright futures carry notional, not premium, so a single futures exit
        # would swamp the ratio. Keep them in the P&L and out of the denominator.
        opt = [e for e in g if e["st"] != "FUT"]
        prem = sum(e["premium"] for e in opt)
        pnl = sum(e["pnl"] for e in g)
        opt_pnl = sum(e["pnl"] for e in opt)
        top = max(g, key=lambda e: e["pnl"])
        out.append({
            "bucket": label, "exits": len(g),
            "premium": round(prem, 2), "pnl": round(pnl, 2),
            "return_on_premium": round(opt_pnl / prem, 4) if prem else None,
            "wins": sum(1 for e in g if e["pnl"] > 0),
            "largest_contributor": top["sym"],
            "pnl_without_largest": round(pnl - top["pnl"], 2),
        })
    return out


def cash_gap(trades, tz, keep_stock):
    """realized_pnl never books the loss on an option that expires worthless.

    Comparing it against true cash movement surfaces that hole, which matters
    because the workbook's totals would otherwise read better than the account.
    """
    rows = [t for t in trades if keep_stock or t["sec_type"] != "STK"]
    cash = sum((1 if t["side"] == "SELL" else -1) * t.get("net_amount", 0.0)
               - t.get("commission", 0.0) for t in rows)
    realized = sum(t.get("realized_pnl", 0.0) for t in rows)
    expiries = [t for t in rows if t["price"] == 0]
    per_sym = collections.defaultdict(float)
    for t in rows:
        per_sym[t["symbol"]] += ((1 if t["side"] == "SELL" else -1) * t.get("net_amount", 0.0)
                                 - t.get("commission", 0.0) - t.get("realized_pnl", 0.0))
    worst = sorted(per_sym.items(), key=lambda kv: kv[1])[:5]
    return {"cash_flow": round(cash, 2), "realized": round(realized, 2),
            "gap": round(cash - realized, 2), "expiry_rows": len(expiries),
            "worst_gaps": [[s, round(v, 2)] for s, v in worst if v < -1]}


# ----------------------------------------------------------------- writing ---

def write_csv(exits, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(["date", "day_he", "time_session", "ticker", "sec_type", "qty",
                    "exit_price", "multiplier", "proceeds_usd", "net_pnl_usd",
                    "premium_paid_usd", "entry_price", "pct_of_premium",
                    "hold_minutes", "commission_usd"])
        for e in exits:
            w.writerow([e["date"], e["day"], e["time"], e["sym"], e["st"], e["qty"],
                        e["px"], e["mult"], e["amt"], e["pnl"], e["premium"],
                        e["entry"], e["ret"], e["hold"], e["com"]])


def write_xlsx(exits, audit, path, notes):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    arial = lambda **k: Font(name="Arial", **k)
    hdr, body, bold = arial(bold=True, size=10, color="FFFFFF"), arial(size=10), arial(bold=True, size=10)
    navy = PatternFill("solid", fgColor="1F3B57")
    alt = PatternFill("solid", fgColor="F2F5F8")
    tot = PatternFill("solid", fgColor="E4EAF0")
    rule = Border(bottom=Side(style="thin", color="D0D7DE"))

    wb = Workbook()
    ws = wb.active
    ws.title = "עסקאות"
    ws.sheet_view.rightToLeft = True
    cols = [("תאריך", 12), ("יום", 9), ("שעה", 9), ("טיקר", 10), ("סוג", 8), ("כמות", 9),
            ("מחיר יציאה", 12), ("מכפיל", 9), ("תמורה ($)", 13), ("תוצאה נטו ($)", 14),
            ("פרמיה ששולמה ($)", 17), ("מחיר כניסה", 12), ("% מהפרמיה", 12),
            ("החזקה (דק׳)", 13), ("עמלה ($)", 11)]
    for i, (h, w) in enumerate(cols, 1):
        c = ws.cell(1, i, h)
        c.font, c.fill = hdr, navy
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    for j, e in enumerate(exits):
        x = j + 2
        for i, v in enumerate([e["date"], e["day"], e["time"], e["sym"], e["st"], e["qty"],
                               e["px"], None, e["amt"], e["pnl"], None, None, None,
                               e["hold"] if e["hold"] != "" else None, e["com"]], 1):
            c = ws.cell(x, i, v)
            c.font, c.border = body, rule
            if j % 2:
                c.fill = alt
        # Derived cells stay formulas so the sheet still adds up after an edit.
        ws.cell(x, 8, '=IF(F{0}*G{0}=0,"",ROUND(I{0}/(F{0}*G{0}),0))'.format(x))
        ws.cell(x, 11, '=I{0}-J{0}'.format(x))
        ws.cell(x, 12, '=IF(OR(F{0}=0,H{0}=""),"",K{0}/(F{0}*H{0}))'.format(x))
        ws.cell(x, 13, '=IF(K{0}=0,"",J{0}/K{0})'.format(x))
        for i, fmt in ((6, '#,##0'), (7, MONEY2), (8, '#,##0'), (9, MONEY), (10, MONEY),
                       (11, MONEY), (12, MONEY2), (13, PCT), (14, '#,##0'), (15, MONEY2)):
            ws.cell(x, i).number_format = fmt

    last, trow = len(exits) + 1, len(exits) + 2
    c = ws.cell(trow, 5, 'סה"כ')
    c.font, c.fill = bold, tot
    for i, f, fmt in ((6, '=SUM(F2:F{})'.format(last), '#,##0'),
                      (9, '=SUM(I2:I{})'.format(last), MONEY),
                      (10, '=SUM(J2:J{})'.format(last), MONEY),
                      (11, '=SUM(K2:K{})'.format(last), MONEY),
                      (13, '=IF(K{0}=0,"",J{0}/K{0})'.format(trow), PCT),
                      (15, '=SUM(O2:O{})'.format(last), MONEY2)):
        c = ws.cell(trow, i, f)
        c.font, c.fill, c.number_format = bold, tot, fmt
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:O{}".format(last)
    for k, txt in enumerate(notes, start=trow + 2):
        ws.cell(k, 1, txt).font = arial(size=9, italic=True)

    # ---- per ticker
    w2 = wb.create_sheet("לפי טיקר")
    w2.sheet_view.rightToLeft = True
    syms = sorted({e["sym"] for e in exits},
                  key=lambda s: -sum(e["pnl"] for e in exits if e["sym"] == s))
    for i, (h, w) in enumerate([("טיקר", 12), ("יציאות", 10), ("זוכות", 9), ("חוזים", 10),
                                ("פרמיה ($)", 14), ("תוצאה נטו ($)", 15), ("% מהפרמיה", 12)], 1):
        c = w2.cell(1, i, h)
        c.font, c.fill = hdr, navy
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        w2.column_dimensions[get_column_letter(i)].width = w
    for j, s in enumerate(syms):
        x = j + 2
        w2.cell(x, 1, s).font = body
        for i, f in ((2, '=COUNTIF(עסקאות!$D$2:$D${1},$A{2})'),
                     (3, '=COUNTIFS(עסקאות!$D$2:$D${1},$A{2},עסקאות!$J$2:$J${1},">0")'),
                     (4, '=SUMIF(עסקאות!$D$2:$D${1},$A{2},עסקאות!$F$2:$F${1})'),
                     (5, '=SUMIF(עסקאות!$D$2:$D${1},$A{2},עסקאות!$K$2:$K${1})'),
                     (6, '=SUMIF(עסקאות!$D$2:$D${1},$A{2},עסקאות!$J$2:$J${1})')):
            w2.cell(x, i, f.format(None, last, x)).font = body
        w2.cell(x, 7, '=IF(E{0}=0,"",F{0}/E{0})'.format(x)).font = body
        for i, fmt in ((2, '#,##0'), (3, '#,##0'), (4, '#,##0'), (5, MONEY), (6, MONEY), (7, PCT)):
            w2.cell(x, i).number_format = fmt
        if j % 2:
            for i in range(1, 8):
                w2.cell(x, i).fill = alt
    t2 = len(syms) + 2
    c = w2.cell(t2, 1, 'סה"כ')
    c.font, c.fill = bold, tot
    for i, f, fmt in ((2, '=SUM(B2:B{})'.format(t2 - 1), '#,##0'),
                      (3, '=SUM(C2:C{})'.format(t2 - 1), '#,##0'),
                      (4, '=SUM(D2:D{})'.format(t2 - 1), '#,##0'),
                      (5, '=SUM(E2:E{})'.format(t2 - 1), MONEY),
                      (6, '=SUM(F2:F{})'.format(t2 - 1), MONEY),
                      (7, '=IF(E{0}=0,"",F{0}/E{0})'.format(t2), PCT)):
        c = w2.cell(t2, i, f)
        c.font, c.fill, c.number_format = bold, tot, fmt
    w2.cell(t2 + 2, 1, "אחוז מהפרמיה בחוזים עתידיים אינו בר-פירוש — שם העלות היא נושׁנל ולא פרמיה."
            ).font = arial(size=9, italic=True)

    # ---- per day
    w3 = wb.create_sheet("לפי יום")
    w3.sheet_view.rightToLeft = True
    days = []
    for e in exits:
        if e["date"] not in [d[0] for d in days]:
            days.append((e["date"], e["day"]))
    for i, (h, w) in enumerate([("תאריך", 13), ("יום", 10), ("יציאות", 10), ("זוכות", 9),
                                ("תוצאה נטו ($)", 15), ("מצטבר ($)", 14)], 1):
        c = w3.cell(1, i, h)
        c.font, c.fill = hdr, navy
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        w3.column_dimensions[get_column_letter(i)].width = w
    for j, (d, dn) in enumerate(days):
        x = j + 2
        w3.cell(x, 1, d).font = body
        w3.cell(x, 2, dn).font = body
        w3.cell(x, 3, '=COUNTIF(עסקאות!$A$2:$A${},$A{})'.format(last, x)).font = body
        w3.cell(x, 4, '=COUNTIFS(עסקאות!$A$2:$A${0},$A{1},עסקאות!$J$2:$J${0},">0")'.format(last, x)).font = body
        w3.cell(x, 5, '=SUMIF(עסקאות!$A$2:$A${0},$A{1},עסקאות!$J$2:$J${0})'.format(last, x)).font = body
        w3.cell(x, 6, '=E2' if j == 0 else '=F{}+E{}'.format(x - 1, x)).font = body
        for i, fmt in ((3, '#,##0'), (4, '#,##0'), (5, MONEY), (6, MONEY)):
            w3.cell(x, i).number_format = fmt
    t3 = len(days) + 2
    c = w3.cell(t3, 2, 'סה"כ')
    c.font, c.fill = bold, tot
    for i, f, fmt in ((3, '=SUM(C2:C{})'.format(t3 - 1), '#,##0'),
                      (4, '=SUM(D2:D{})'.format(t3 - 1), '#,##0'),
                      (5, '=SUM(E2:E{})'.format(t3 - 1), MONEY)):
        c = w3.cell(t3, i, f)
        c.font, c.fill, c.number_format = bold, tot, fmt

    # ---- audit
    w4 = wb.create_sheet("סינון")
    w4.sheet_view.rightToLeft = True
    for i, (h, w) in enumerate([("שלב", 54), ("שורות", 11), ("השפעה על התוצאה ($)", 22)], 1):
        c = w4.cell(1, i, h)
        c.font, c.fill = hdr, navy
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        w4.column_dimensions[get_column_letter(i)].width = w
    for j, (label, n, pnl) in enumerate(audit):
        x = j + 2
        w4.cell(x, 1, label).font = body
        w4.cell(x, 2, n).font = body
        w4.cell(x, 3, round(pnl, 2)).font = body
        w4.cell(x, 2).number_format = '#,##0'
        w4.cell(x, 3).number_format = MONEY
        if j % 2:
            for i in range(1, 4):
                w4.cell(x, i).fill = alt
    t4 = len(audit) + 2
    c = w4.cell(t4, 1, '= הסט המסונן (גיליון "עסקאות")')
    c.font, c.fill = bold, tot
    for i, f in ((2, '=SUM(B2:B{})'.format(t4 - 1)), (3, '=SUM(C2:C{})'.format(t4 - 1))):
        c = w4.cell(t4, i, f)
        c.font, c.fill = bold, tot
        c.number_format = '#,##0' if i == 2 else MONEY
    w4.cell(t4 + 1, 1, "בדיקת התאמה מול גיליון העסקאות").font = arial(size=10, italic=True)
    c = w4.cell(t4 + 1, 3, "='עסקאות'!J{}".format(trow))
    c.font, c.number_format = arial(size=10, italic=True), MONEY

    wb.save(path)


# -------------------------------------------------------------------- main ---

def main():
    p = argparse.ArgumentParser()
    p.add_argument("trades_json")
    p.add_argument("outdir")
    p.add_argument("--week-start")
    p.add_argument("--tickers", type=lambda s: [x.strip().upper() for x in s.split(",") if x.strip()])
    p.add_argument("--drop-carry", type=lambda s: [x.strip().upper() for x in s.split(",") if x.strip()])
    p.add_argument("--carry-day")
    p.add_argument("--keep-stock", action="store_true")
    p.add_argument("--keep-buys", action="store_true")
    p.add_argument("--tz-offset", type=int, default=-4)
    args = p.parse_args()
    if args.drop_carry and not args.carry_day:
        sys.exit("--drop-carry needs --carry-day")

    trades = load_trades(args.trades_json)
    if not args.week_start:
        # Pick the week holding the most trades rather than the latest one: a
        # Friday-evening futures fill rolls into the next Monday, and that lone
        # straggler would otherwise be read as "the week" and empty the report.
        weeks = collections.Counter()
        for t in trades:
            d = session_day(session_time(t["trade_time"], args.tz_offset), t["price"] == 0)
            weeks[d - datetime.timedelta(days=d.weekday())] += 1
        if not weeks:
            sys.exit("no trades in the file")
        args.week_start = str(weeks.most_common(1)[0][0])

    holds = holding_times(trades, args.tz_offset)
    kept, audit = build(trades, holds, args)
    exits = to_exits(kept, args.tz_offset, holds)

    # The expiry line was reserved in build() before the charge was known; fill it in
    # so the audit chain still reconciles against the exits it produced.
    expired = [e for e in exits if e.get("expired")]
    charge = sum(e["pnl"] for e in expired)
    for i, (label, n, _) in enumerate(audit):
        if label.startswith("(=) שורות פקיעה"):
            audit[i] = (label, len(expired), charge)
    if not exits:
        sys.exit("every trade was filtered out — check --week-start and --tickers")

    os.makedirs(args.outdir, exist_ok=True)
    xlsx = os.path.join(args.outdir, "trades-filtered.xlsx")
    csvp = os.path.join(args.outdir, "trades-filtered.csv")
    notes = [
        "פרמיה ששולמה = תמורה פחות תוצאה נטו. מחיר כניסה = פרמיה / (כמות × מכפיל).",
        "התוצאה נטו כפי שמדווחת ב-IBKR כבר מנוכה עמלות משתי רגלי הסיבוב.",
        "מקור: Interactive Brokers, get_account_trades. שעות בשעון המסחר.",
    ]
    write_csv(exits, csvp)
    write_xlsx(exits, audit, xlsx, notes)

    wins = [e for e in exits if e["pnl"] > 0]
    losses = [e for e in exits if e["pnl"] < 0]
    opt = [e for e in exits if e["st"] != "FUT"]
    opt_prem = sum(e["premium"] for e in opt)
    net = sum(e["pnl"] for e in exits)
    chain = audit[0][2] + sum(a[2] for a in audit[1:])

    summary = {
        "week_start": args.week_start,
        "exits": len(exits),
        "fills_kept": len(kept),
        "tickers": sorted({e["sym"] for e in exits}),
        "net_pnl": round(net, 2),
        "wins": len(wins), "losses": len(losses),
        "win_rate": round(len(wins) / len(exits), 4),
        "profit_factor": (round(sum(e["pnl"] for e in wins) / -sum(e["pnl"] for e in losses), 3)
                          if losses else None),
        "commission_sell_leg": round(sum(e["com"] for e in exits), 2),
        "option_premium": round(opt_prem, 2),
        "return_on_option_premium": (round(sum(e["pnl"] for e in opt) / opt_prem, 4)
                                     if opt_prem else None),
        "by_day": [[str(d), round(sum(e["pnl"] for e in exits if e["date"] == d), 2)]
                   for d in sorted({e["date"] for e in exits})],
        "by_ticker": sorted([[s, round(sum(e["pnl"] for e in exits if e["sym"] == s), 2)]
                             for s in {e["sym"] for e in exits}], key=lambda r: -r[1]),
        "by_hold": hold_analysis(exits),
        "expiries": {
            "count": len(expired),
            "contracts": sum(e["qty"] for e in expired),
            "premium_lost": round(-charge, 2),
            "share_of_gross_loss": (
                round(-charge / -sum(e["pnl"] for e in exits if e["pnl"] < 0), 4)
                if any(e["pnl"] < 0 for e in exits) else None),
            "by_ticker": sorted(
                [[s_, round(-sum(e["pnl"] for e in expired if e["sym"] == s_), 2)]
                 for s_ in {e["sym"] for e in expired}], key=lambda r: -r[1]),
        },
        "audit": [[a, n, round(p, 2)] for a, n, p in audit],
        "audit_reconciles": abs(chain - net) < 0.05,
        "cash_check": cash_gap(trades, args.tz_offset, args.keep_stock),
        "files": {"xlsx": xlsx, "csv": csvp},
    }
    with open(os.path.join(args.outdir, "summary.json"), "w", encoding="utf-8") as fh:
        json.dump(summary, fh, ensure_ascii=False, indent=2)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
